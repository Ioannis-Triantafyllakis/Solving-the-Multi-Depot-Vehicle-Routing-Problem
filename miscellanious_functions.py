import matplotlib.pyplot as plt
import random
import statistics
import pandas as pd
from openpyxl import Workbook
import time
from MDVRP_KMeansFunc import *
from ClarkeAndWrightSavingsAlgorithm import *
from MDVRP_HierClustFunc import HierClusteringBasedNodeSelection

def MDVRPModelInstances(NoOfCustomers, Seed, NoOfVehicles, NoOfDepots, Grid, VehicleCap, MaximumDem):
    '''
    This function takes as input: 1) The number of customers, 2) a Seed, 3) The number of vehicles, 4) The number of depots, 
    5) A grid, which is the maximum value a customer's coordinates can get, and the length of x and y axis at the same time, 
    6) The Vehicle Capacity, and 7) The maximum demand a customer can have. By using a Seed number with the "random" framework, 
    reproducible MDVRP instances are created in the form of a dictionary.
    '''
    random.seed(Seed)

    AllDepots = []
    for i in range(1, int(NoOfDepots)+1):
        dep = ["0"+str(i), random.randint(1, Grid), random.randint(1, Grid)]        
        AllDepots.append(dep)
    
    all_Customers = [] # Lists with customers
    all_Nodes = []     # List with all Nodes (customers + depots)

    for i in AllDepots: # Appending the depots to the list of lists that includes all nodes
        all_Nodes.append(i)
    
    for i in range(1, int(NoOfCustomers)+1):
        cust = []
        cust = [i, random.randint(1, Grid), random.randint(1, Grid)]
        all_Customers.append(cust)
        all_Nodes.append(cust)

    all_ids = []
    all_x = []
    all_y = []
    all_custs_ids = []
    all_dem = []

    for i in range(0, len(all_Customers)):
        dem = [ all_Customers[i][0], random.randint(1, MaximumDem)]
        all_dem.append(dem)


    for i in range(0, len(all_Nodes)):
        id = all_Nodes[i][0]
        all_ids.append(id)
        x = all_Nodes[i][1]
        all_x.append(x)
        y = all_Nodes[i][2]
        all_y.append(y)
        if i==0:
            continue
        else:
            cust_id = all_Nodes[i][0]
            all_custs_ids.append(cust_id)

    DictOfVehiclesRoutes = {}
    for i in range(1, int(NoOfVehicles)+1):
        DictOfVehiclesRoutes['RouteOfVehicle_%s' % i] = []
        
    return {"allCustomers": all_Customers, 
            "allNodes": all_Nodes, 
            "allIds": all_ids, 
            "allxs": all_x, 
            "allys": all_y, 
            "NoOfCusts":NoOfCustomers,
            "AllDepots": AllDepots,
            "NoOfVehicles": NoOfVehicles,
            "all_dem": all_dem,
            "VehicleCap": VehicleCap, 
            "Seed": Seed}
    # all_Customers and All_Nodes are lists of lists. Format is [id, x, y].    


def DepotsAndNodesPairsLists(SelectionDict, AllNodesList):
    '''
    This function creates a list of lists of lists, where each sublist corresponds to a node with a [id, x coord, y coord]
    format, the second level lists are a depot with the nodes assigned to it, and the main list is a "container" list.
    Inputs:
    1) Dictionary of depots-nodes's pairs (created by the "selection" phase)
    2) A list of lists of all existing nodes' info with a [id, x coord, y coord] format
    The list it returns contains other lists, that are the input for distinct VRP's
    '''
    ListOfAllDepotsNodesPairs = list()

    for key, value in SelectionDict.items():
        DepotNodesPair = list()
        for i in AllNodesList:
            if i[0] in value or i[0] == key:
                DepotNodesPair.append([i[0], i[1], i[2]])
        ListOfAllDepotsNodesPairs.append(DepotNodesPair)

    return ListOfAllDepotsNodesPairs


def RoutingInfoContainer(ListOfPairs, RoutingHeuristic, NoOfVehicles, Cap, all_dem):
    '''
    This function gets as input: 1) The list with all nodes of each distinct VRP, 2) The routing heuristic function, 3)
    The number of vehicles, 4) the total capacity of each vehicle, and 5), the list with each customer's demand.
    This function works as "container" that stores the results of all distinct VRP's solved.
    '''
    RoutesContainer = []
    CostsContainer = []
    TotalCostsContainer = []
    DistMatricesContainer = []

    for list in ListOfPairs:        
        RoutesContainer.append(RoutingHeuristic(list, Cap, all_dem, NoOfVehicles)["Routes"])
        CostsContainer.append(RoutingHeuristic(list, Cap, all_dem, NoOfVehicles)["AllCosts"])
        TotalCostsContainer.append(RoutingHeuristic(list, Cap, all_dem, NoOfVehicles)["TotalCost"])
        DistMatricesContainer.append(RoutingHeuristic(list, Cap, all_dem, NoOfVehicles)["Distance_Matrix"])

    return {"RoutesContainer":RoutesContainer,
            "CostsContainer":CostsContainer,
            "TotalCostsContainer":TotalCostsContainer, 
            "DistMatricesContainer":DistMatricesContainer}


def SolutionPlot(all_xs, all_ys, all_Ids, RoutesContainer, GridSize):
    '''
    This function creates a "matplotlib" figure with the MDVRP solution.
    '''
    fig, ax = plt.subplots()

    for route in RoutesContainer:
        for subroute in route:
            route_xcoords = list()
            route_ycoords = list()
            for node in subroute:
                idx = all_Ids.index(node)
                xcoord = all_xs[idx]
                ycoord = all_ys[idx]
                route_xcoords.append(xcoord)
                route_ycoords.append(ycoord)
            ax.plot(route_xcoords, route_ycoords, '-o')

    for xi, yi, pidi in zip(all_xs, all_ys, all_Ids):
            ax.annotate(str(pidi), xy=(xi,yi), fontsize = 13)

    # Show the plot.
    fig.set_size_inches(8, 8)

    plt.xlim([0, GridSize])
    plt.ylim([0, GridSize])
    plt.style.use("default") # seaborn, default, seaborn-pastel

    return fig


def SolutionMetricsFinder(CostsContainer, TotalCostsContainer, all_dem, RoutesContainer, DictOfDepotsAndNodesPairs):
    '''
    This function gets results from the distinct VRP's solved, and calculates and returns 9 metrics related to the MDVRP solution.
    '''
    AllRoutesCost = list()
    for RouteList in CostsContainer:
        for RouteCost in RouteList:
            AllRoutesCost.append(RouteCost)

    DemandSatisfied = 0
    NoOfCustsVisited = 0
    for route in RoutesContainer:
        for subroute in route:
            for r in subroute:
                if subroute[-1] == r or subroute[0] == r:
                    continue
                else:
                    NoOfCustsVisited = NoOfCustsVisited + 1
                    for cust in all_dem:
                        if r == cust[0]:
                            DemandSatisfied = DemandSatisfied + cust[1]

    TotalDemand = 0
    NoOfCustsToBeVisited = 0
    for k, v in DictOfDepotsAndNodesPairs.items():
        for node in v:
            NoOfCustsToBeVisited = NoOfCustsToBeVisited + 1
            for i in all_dem:
                if node == i[0]:
                    TotalDemand = TotalDemand + i[1]

    PctOfTotalDemandSatisfied = round((DemandSatisfied/TotalDemand)*100, 2)
    PctOfCustomersVisited = round((NoOfCustsVisited/NoOfCustsToBeVisited)*100, 2)

    AvgRouteCost = round(sum(AllRoutesCost)/len(AllRoutesCost), 2)
    MedianRouteCost = round(statistics.median(AllRoutesCost), 2) 

    AvgDepotCost = round(sum(TotalCostsContainer)/len(TotalCostsContainer), 2)
    MedianDepotCost = round(statistics.median(TotalCostsContainer), 2)
    TotalCost = round(sum(TotalCostsContainer), 2)

    return {"TotalCost" : TotalCost,
            "AvgRouteCost" : AvgRouteCost,
            "MedianRouteCost" : MedianRouteCost,
            "AvgDepotCost" : AvgDepotCost,
            "MedianDepotCost" : MedianDepotCost,
            "PctOfTotalDemandSatisfied" : PctOfTotalDemandSatisfied,
            "DemandSatisfied" : DemandSatisfied,
            "TotalDemand" : TotalDemand, 
            "PctOfCustomersVisited" : PctOfCustomersVisited}


def MDVRP_BenchmarkInstances(Instance, Seed):

    with open(Instance) as f:
        lines = f.readlines()

    df = pd.DataFrame(columns=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11])

    InstanceList = list()
    for item in lines:
        InstanceList.append(list(item.split()))
    df = pd.DataFrame(InstanceList)

    NoOfVehicles = int(df[1][0])
    NoOfCusts = int(df[2][0])
    NoOfDepots = int(df[3][0])

    VehicleCap = int(df[1][1])

    all_Customers = list()
    all_Nodes = list()
    allIds = list()
    all_dem = list()
    allxs = list()
    allys = list()
    allDepots = list()

    for row in range(NoOfVehicles+1, NoOfCusts+NoOfVehicles+1):
        cust = df[row:row+1]

        cust_id = int(cust[0].astype("int"))
        cust_x = int(cust[1].astype("int"))
        cust_y = int(cust[2].astype("int"))
        cust_dem = int(cust[4].astype("int"))

        customer = [cust_id, cust_x, cust_y]
        demand = [cust_id, cust_dem]
    
        all_Customers.append(customer)
        all_dem.append(demand)

    counter = 1
    for row in range(NoOfVehicles+1, NoOfCusts+NoOfVehicles+NoOfDepots+1):
        node = df[row:row+1]
        node_id = int(node[0].astype("int"))
        if node_id > NoOfCusts:
            node_id = "0" + str(counter)
            counter = counter + 1
            allIds.append(node_id)
            node_x = int(node[1].astype("int"))
            node_y = int(node[2].astype("int"))

            allxs.append(node_x)
            allys.append(node_y)
            #all_Nodes.append([node_id, node_x, node_y])
            all_Nodes.insert(0, [node_id, node_x, node_y])
            depot = [node_id, node_x, node_y]
            allDepots.append(depot)

        else:
            node_x = int(node[1].astype("int"))
            node_y = int(node[2].astype("int"))
            allxs.append(node_x)
            allys.append(node_y)
            all_Nodes.append([node_id, node_x, node_y])
            allIds.append(node_id)

    return {"allCustomers": all_Customers, 
            "allNodes": all_Nodes, 
            "allIds": allIds, 
            "allxs": allxs, 
            "allys": allys, 
            "NoOfCusts":NoOfCusts,
            "AllDepots": allDepots,
            "NoOfVehicles": NoOfVehicles,
            "all_dem": all_dem,
            "VehicleCap": VehicleCap, 
            "Seed": Seed}



def BatchTestingExcelWriter(NoOfMetrics, FromSeed, ToSeed, noc, nov, nod, nog, nocap, nodem):

    workbook = Workbook()
    sheet = workbook.active

    NoOfLines = NoOfMetrics
    Counter = 0
    Simulations = ToSeed - FromSeed

    # 1st heuristic
    for i in range(Counter+1, Counter+2): 
        for j in range(1, Simulations+2):                                           
            
            st = time.time()
            cpu_st = time.process_time()
            inst = MDVRPModelInstances(noc, j, nov, nod, nog, nocap, nodem) 

            DictOfDepotsAndNodesPairs = KMeansClusteringBasedNodesSelection(inst) # "KMeans-Clustering-Based Clarke & Wright Heuristic"
            ListsOfPairs = DepotsAndNodesPairsLists(DictOfDepotsAndNodesPairs, inst["allNodes"])
            xx = inst["allxs"]        
            yy = inst["allys"]

            all_ids = inst["allIds"]
            NoOfVehicles = inst["NoOfVehicles"]
            Cap = inst["VehicleCap"]
            all_dem = inst["all_dem"]

            Container = RoutingInfoContainer(ListsOfPairs, ClarkeAndWrightSavingsAlgorithmWithVehConstraint, NoOfVehicles, Cap, all_dem)

            Metrics = SolutionMetricsFinder(Container["CostsContainer"], Container["TotalCostsContainer"], all_dem, Container["RoutesContainer"], DictOfDepotsAndNodesPairs)
            
            et = time.time()
            cpu_et = time.process_time()

            elapsed_time = round(et - st, 3)
            cpu_elapsed_time = round((cpu_et - cpu_st)/100, 3)

            sheet.cell(row=i, column=j).value = Metrics["TotalCost"]
            sheet.cell(row=i+1, column=j).value = Metrics["AvgRouteCost"]
            sheet.cell(row=i+2, column=j).value = Metrics["MedianRouteCost"]
            sheet.cell(row=i+3, column=j).value = Metrics["AvgDepotCost"]
            sheet.cell(row=i+4, column=j).value = Metrics["MedianDepotCost"]
            sheet.cell(row=i+5, column=j).value = Metrics["PctOfTotalDemandSatisfied"]
            sheet.cell(row=i+6, column=j).value = Metrics["PctOfCustomersVisited"]
            sheet.cell(row=i+7, column=j).value = elapsed_time
            sheet.cell(row=i+8, column=j).value = cpu_elapsed_time
            
            Counter = Counter + NoOfMetrics + 2
            NoOfLines = Counter + NoOfMetrics 

    # 2nd heursitic
    Counter = 0
    for i in range(Counter+1, Counter+2): 
        for j in range(1, Simulations+2):                                         
            
            st = time.time()
            cpu_st = time.process_time()
            inst = MDVRPModelInstances(noc, j, nov, nod, nog, nocap, nodem) 

            DictOfDepotsAndNodesPairs = HierClusteringBasedNodeSelection(inst, "ward")
            ListsOfPairs = DepotsAndNodesPairsLists(DictOfDepotsAndNodesPairs, inst["allNodes"])
            xx = inst["allxs"]        
            yy = inst["allys"]

            all_ids = inst["allIds"]
            NoOfVehicles = inst["NoOfVehicles"]
            Cap = inst["VehicleCap"]
            all_dem = inst["all_dem"]

            Container = RoutingInfoContainer(ListsOfPairs, ClarkeAndWrightSavingsAlgorithmWithVehConstraint, NoOfVehicles, Cap, all_dem)

            Metrics = SolutionMetricsFinder(Container["CostsContainer"], Container["TotalCostsContainer"], all_dem, Container["RoutesContainer"], DictOfDepotsAndNodesPairs)
            
            et = time.time()
            cpu_et = time.process_time()

            elapsed_time = round(et - st, 3)
            cpu_elapsed_time = round((cpu_et - cpu_st)/100, 3)

            sheet.cell(row=i+10, column=j).value = Metrics["TotalCost"]
            sheet.cell(row=i+11, column=j).value = Metrics["AvgRouteCost"]
            sheet.cell(row=i+12, column=j).value = Metrics["MedianRouteCost"]
            sheet.cell(row=i+13, column=j).value = Metrics["AvgDepotCost"]
            sheet.cell(row=i+14, column=j).value = Metrics["MedianDepotCost"]
            sheet.cell(row=i+15, column=j).value = Metrics["PctOfTotalDemandSatisfied"]
            sheet.cell(row=i+16, column=j).value = Metrics["PctOfCustomersVisited"]
            sheet.cell(row=i+17, column=j).value = elapsed_time
            sheet.cell(row=i+18, column=j).value = cpu_elapsed_time
            
            Counter = Counter + NoOfMetrics + 2
            NoOfLines = Counter + NoOfMetrics

    # 3rd heursitic
    Counter = 0
    for i in range(Counter+1, Counter+2): 
        for j in range(1, Simulations+2):                                         
            
            st = time.time()
            cpu_st = time.process_time()
            inst = MDVRPModelInstances(noc, j, nov, nod, nog, nocap, nodem) 

            DictOfDepotsAndNodesPairs = HierClusteringBasedNodeSelection(inst, "complete")
            ListsOfPairs = DepotsAndNodesPairsLists(DictOfDepotsAndNodesPairs, inst["allNodes"])
            xx = inst["allxs"]        
            yy = inst["allys"]

            all_ids = inst["allIds"]
            NoOfVehicles = inst["NoOfVehicles"]
            Cap = inst["VehicleCap"]
            all_dem = inst["all_dem"]

            Container = RoutingInfoContainer(ListsOfPairs, ClarkeAndWrightSavingsAlgorithmWithVehConstraint, NoOfVehicles, Cap, all_dem)

            Metrics = SolutionMetricsFinder(Container["CostsContainer"], Container["TotalCostsContainer"], all_dem, Container["RoutesContainer"], DictOfDepotsAndNodesPairs)
            
            et = time.time()
            cpu_et = time.process_time()

            elapsed_time = round(et - st, 3)
            cpu_elapsed_time = round((cpu_et - cpu_st)/100, 3)

            sheet.cell(row=i+20, column=j).value = Metrics["TotalCost"]
            sheet.cell(row=i+21, column=j).value = Metrics["AvgRouteCost"]
            sheet.cell(row=i+22, column=j).value = Metrics["MedianRouteCost"]
            sheet.cell(row=i+23, column=j).value = Metrics["AvgDepotCost"]
            sheet.cell(row=i+24, column=j).value = Metrics["MedianDepotCost"]
            sheet.cell(row=i+25, column=j).value = Metrics["PctOfTotalDemandSatisfied"]
            sheet.cell(row=i+26, column=j).value = Metrics["PctOfCustomersVisited"]
            sheet.cell(row=i+27, column=j).value = elapsed_time
            sheet.cell(row=i+28, column=j).value = cpu_elapsed_time
            
            Counter = Counter + NoOfMetrics + 2
            NoOfLines = Counter + NoOfMetrics 

    # 4th heursitic
    Counter = 0
    for i in range(Counter+1, Counter+2): 
        for j in range(1, Simulations+2):                                         
            
            st = time.time()
            cpu_st = time.process_time()
            inst = MDVRPModelInstances(noc, j, nov, nod, nog, nocap, nodem) 

            DictOfDepotsAndNodesPairs = HierClusteringBasedNodeSelection(inst, "average")
            ListsOfPairs = DepotsAndNodesPairsLists(DictOfDepotsAndNodesPairs, inst["allNodes"])
            xx = inst["allxs"]        
            yy = inst["allys"]

            all_ids = inst["allIds"]
            NoOfVehicles = inst["NoOfVehicles"]
            Cap = inst["VehicleCap"]
            all_dem = inst["all_dem"]

            Container = RoutingInfoContainer(ListsOfPairs, ClarkeAndWrightSavingsAlgorithmWithVehConstraint, NoOfVehicles, Cap, all_dem)

            Metrics = SolutionMetricsFinder(Container["CostsContainer"], Container["TotalCostsContainer"], all_dem, Container["RoutesContainer"], DictOfDepotsAndNodesPairs)
            
            et = time.time()
            cpu_et = time.process_time()

            elapsed_time = round(et - st, 3)
            cpu_elapsed_time = round((cpu_et - cpu_st)/100, 3)

            sheet.cell(row=i+30, column=j).value = Metrics["TotalCost"]
            sheet.cell(row=i+31, column=j).value = Metrics["AvgRouteCost"]
            sheet.cell(row=i+32, column=j).value = Metrics["MedianRouteCost"]
            sheet.cell(row=i+33, column=j).value = Metrics["AvgDepotCost"]
            sheet.cell(row=i+34, column=j).value = Metrics["MedianDepotCost"]
            sheet.cell(row=i+35, column=j).value = Metrics["PctOfTotalDemandSatisfied"]
            sheet.cell(row=i+36, column=j).value = Metrics["PctOfCustomersVisited"]
            sheet.cell(row=i+37, column=j).value = elapsed_time
            sheet.cell(row=i+38, column=j).value = cpu_elapsed_time
            
            Counter = Counter + NoOfMetrics + 2
            NoOfLines = Counter + NoOfMetrics

    ExcelFileName = "Sims-From-Seed-"+str(FromSeed)+"-to-"+str(ToSeed)+"-"+str(noc)+"-"+str(nov)+"-"+str(nod)+"-"+str(nog)+"-"+str(nocap)+"-"+str(nodem)+".xlsx"
    
    workbook.save(filename=ExcelFileName)


def CostFinder(Route, Matrix):
    Cost = 0
    for i in Route[:-1]:
        Cost = Cost + Matrix[Route.index(i)][Route.index(i)+1]
    return Cost